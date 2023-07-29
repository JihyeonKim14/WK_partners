import pandas as pd
import slack
import os
from slack import WebClient
from slack.errors import SlackApiError
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_and_send_file_to_slack():
    id = "#########"
    pw = "##########"
    host = "##########"
    port = "#########"
    db = "############c"

    engine = create_engine(f"mysql+pymysql://{id}:{pw}@{host}:{port}/{db}")
    categories = ["한식", "기타 외국식", "기타 외식", "분식", "서양식", "아이스크림 빙수", "음료 (커피 외)", "일식", "제과제빵", "주점", "중식", "치킨", "커피", "패스트푸드", "피자"]
    
    main_query1 = f"""
        SELECT *
        FROM ftc.FTC_LONGLIST_RESULT_2021 AS A
        """
    df1 = pd.read_sql(main_query1, engine)
    df1 = df1[df1['업종'].isin(categories)]

    main_query2 = f"""
        SELECT *
        FROM ftc.FTC_LONGLIST_RESULT_2022 AS A
        """
    df2 = pd.read_sql(main_query2, engine)
    df2 = df2[df2['업종'].isin(categories)]
    main_query3 = f"""
        SELECT t1.*
        FROM ftc.FTC_LONGLIST_RESULT_2022 t1
        WHERE t1.search_dt = CURDATE()
        AND t1.page_id NOT IN ( SELECT t2.page_id FROM ftc.FTC_LONGLIST_RESULT_2022 t2
        WHERE t2.search_dt = DATE_SUB(CURDATE(), INTERVAL 1 WEEK) )
        """
    df3 = pd.read_sql(main_query3, engine)
    df3 = df3[df3['업종'].isin(categories)]
    main_query4 = f"""
        SELECT t1.*
        FROM ftc.FTC_LONGLIST_RESULT_2022 t1
        WHERE t1.search_dt = DATE_SUB(CURDATE(), INTERVAL 1 WEEK)
        AND t1.page_id NOT IN ( SELECT t2.page_id FROM ftc.FTC_LONGLIST_RESULT_2022 t2
        WHERE t2.search_dt = CURDATE() )
        """
    
    df4 = pd.read_sql(main_query4, engine)
    df4 = df4[df4['업종'].isin(categories)]
    date = datetime.today().strftime("%Y-%m-%d")
    # Create an Excel writer using openpyxl
    
    df1 = df1[df1['search_dt']== date]
    df2 = df2[df2['search_dt']== date]
    
    writer = pd.ExcelWriter(f'공정위_롱리스트_{date}.xlsx', engine='openpyxl')

    # Convert the DataFrames to Excel sheets
    df1.to_excel(writer, index=False, sheet_name='2021년')
    df2.to_excel(writer, index=False, sheet_name='2022년')
    df3.to_excel(writer, index=False, sheet_name='주 단위 신규 추출 브랜드')
    df4.to_excel(writer, index=False, sheet_name='주 단위 로직 탈락 브랜드')
    
    # Get the workbook
    # workbook = writer.book

    # Set the fill color for the first row
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Iterate over each sheet
    for sheet_name in writer.sheets:
        sheet = writer.sheets[sheet_name]

        # Get the first row range
        first_row_range = sheet['1']

        # Apply the fill color to the first row
        for cell in first_row_range:
            cell.fill = fill

        # Adjust column widths based on content
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
        
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # Save the Excel file
    writer.close()

    # Get the current date
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Create the initial comment indicating that the file has been sent
    initial_comment = f"{current_date} 기준 공정위 롱리스트 산출이 완료되었습니다."


    # Provide a link to the file in the Slack message

    slack_token = '###############################################'
    client = slack.WebClient(token=slack_token)
    response = client.files_upload(
        #channels='# notice',
        channels = '# ds-ai-팀',
        file= f'공정위_롱리스트_{date}.xlsx',
        initial_comment= initial_comment
    )

    # Print the response to the console
    print(response)

# Call the function to generate and send the file to Slack



if __name__ == '__main__':
    generate_and_send_file_to_slack()
